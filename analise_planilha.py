import pandas as pd
import openpyxl
import sys

def analisar_planilha(caminho_arquivo):
    try:
        # Carregar o arquivo Excel
        workbook = openpyxl.load_workbook(caminho_arquivo)
        
        print("=== ANÁLISE DA PLANILHA DE MEMÓRIA DE CÁLCULO ===\n")
        
        # Listar todas as abas
        print("Abas encontradas:")
        for i, sheet_name in enumerate(workbook.sheetnames, 1):
            print(f"{i}. {sheet_name}")
        print()
        
        # Analisar cada aba
        for sheet_name in workbook.sheetnames:
            print(f"=== ABA: {sheet_name} ===")
            
            # Usar pandas para análise mais detalhada
            try:
                df = pd.read_excel(caminho_arquivo, sheet_name=sheet_name)
                
                print(f"Dimensões: {df.shape[0]} linhas x {df.shape[1]} colunas")
                print(f"Colunas: {list(df.columns)}")
                
                # Mostrar primeiras linhas não vazias
                df_clean = df.dropna(how='all')
                if not df_clean.empty:
                    print("\nPrimeiras linhas com dados:")
                    print(df_clean.head(10).to_string())
                
                # Verificar se há fórmulas ou cálculos
                print(f"\nLinhas com dados: {len(df_clean)}")
                
                # Procurar por valores numéricos
                numeric_cols = df.select_dtypes(include=['number']).columns
                if len(numeric_cols) > 0:
                    print(f"Colunas numéricas: {list(numeric_cols)}")
                    for col in numeric_cols:
                        valores_nao_nulos = df[col].dropna()
                        if len(valores_nao_nulos) > 0:
                            print(f"  {col}: min={valores_nao_nulos.min():.2f}, max={valores_nao_nulos.max():.2f}, média={valores_nao_nulos.mean():.2f}")
                
            except Exception as e:
                print(f"Erro ao analisar com pandas: {e}")
                
                # Fallback: usar openpyxl diretamente
                worksheet = workbook[sheet_name]
                print(f"Dimensões: {worksheet.max_row} linhas x {worksheet.max_column} colunas")
                
                # Mostrar algumas células com conteúdo
                print("\nConteúdo das primeiras células:")
                for row in range(1, min(11, worksheet.max_row + 1)):
                    row_data = []
                    for col in range(1, min(6, worksheet.max_column + 1)):
                        cell_value = worksheet.cell(row=row, column=col).value
                        if cell_value is not None:
                            row_data.append(str(cell_value)[:50])  # Limitar tamanho
                        else:
                            row_data.append("")
                    if any(row_data):  # Só mostrar se a linha tem conteúdo
                        print(f"Linha {row}: {' | '.join(row_data)}")
            
            print("\n" + "="*50 + "\n")
        
        workbook.close()
        
    except Exception as e:
        print(f"Erro ao analisar planilha: {e}")
        return False
    
    return True

if __name__ == "__main__":
    caminho = "/home/ubuntu/upload/memoria31-07rev01.xlsx"
    analisar_planilha(caminho)

